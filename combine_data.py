# Intel Confidential

# Ronny Z. Valtonen

from hashlib import new
from openpyxl import *
import openpyxl
from openpyxl.styles import *
from openpyxl.cell import *
from openpyxl.utils import *
from openpyxl.worksheet.dimensions import *
from os.path import exists

#####################################
# BASIC PREP                        #
# pip install openpyxl (3.0.10)     #
# pip install pillow     (3.0.3)    #
# python 3.9.12                     #
# Linux: sudo pacman -S tk          #
# MacOS: brew install python-tk     #
#####################################


def template(cd_data, cd):

        # Open the benchmakrs excel sheet, but make sure it exists.
    path = 'benchmarks.xlsx'

    # Get the worksheets to loop through
    wb = load_workbook(path)
    file_exists = exists(path)
    workbook = load_workbook(filename=path)

    sh = cd.active

    for sheet_name in workbook.sheetnames:
        if file_exists == True:
            workbook.active = workbook[sheet_name]
            second_col = workbook.active['B']
            new_sheet = workbook.active.cell(row=2,column=1)


        # Crossmark
        if new_sheet.value == 'Productivity':
            print("Crossmark data detected, compiling information")

            cross_Data = []

            # Loop through the column and get the values.
            for x in range(len(second_col)):
                # Grab the row column value
                this = second_col[x].value.strip()

                # Add to our tuple, use for appending
                cross_Data += (this, )  


            # Move to the next column
            check_data_b = cd_data['B']
            check_data_c = cd_data['C']
            check_data_d = cd_data['D']

            index = 0

            # Column for Run 1 writing
            if check_data_b[20].value == None:
                print("Populating Run1")
                for p in cross_Data:
                    if index <= 3:
                        index += 1
                        cd_data.append({'B': cross_Data[index-1]})
                print("Moving data to Run 1")
                cd_data.move_range("B25:B28", rows= -5, cols=0)
                #cd.save("combined_Data.xlsx")


            # Column for Run 1 is populated
            if check_data_b[20].value != None:

                # Column for Run 2 writing
                if check_data_c[20].value == None:
                    print("Populating Run2")
                    for r2 in cross_Data:
                        if index <= 3:
                            index += 1
                            cd_data.append({'C': cross_Data[index-1]})
                    print("Moving data to Run 2")
                    cd_data.move_range("C25:C28", rows= -5, cols=0)
                    #cd.save("combined_Data.xlsx")

                # Column for Run 3 writing
                if check_data_c[20].value != None:
                    print("Populating Run 3")
                    for r3 in cross_Data:
                        if index <= 3:
                            index += 1
                            cd_data.append({'D': cross_Data[index-1]})
                    print("Moving data to Run 3")
                    cd_data.move_range("D25:D28", rows=-5, cols=0)
                cd.save("combined_Data.xlsx")



        # PCMark10
        if new_sheet.value == 'Essentials':
            print("PCMark10 data detected, compiling information")

            pc_Data = []
            # Loop through the column and get the values.

            for y in range(len(second_col)):
                # Grab the row column values
                that = second_col[y].value
                
                # Add to our tuple, use for appending
                pc_Data += (that, )

            # Move to the next column
            check_data_b = cd_data['B']
            check_data_c = cd_data['C']
            check_data_d = cd_data['D']     

            pc_index = 0

            # Column for Run 1 Writing
            if check_data_b[6].value == None:   
                print("Populating Run 1")
                for r1 in pc_Data:
                    if pc_index <= 11:
                        pc_index += 1
                        cd_data.append({'B': pc_Data[pc_index-1]})
                print("Moving data to Run 1")
                cd_data.move_range("B25:B36", rows= -19, cols=0)

                c1 = sh['B']
                if c1[37].value != None:
                    cd_data.move_range("B37:B40", rows= -31, cols=0)
                #cd.save("combined_Data.xlsx")

            # Column for Run 1 is populated
            if check_data_b[6].value != None:
                
                # Column for Run 2 writing
                if check_data_c[6].value == None:
                    print("Populated Run2")
                    for pc_r2 in pc_Data:
                        if pc_index <= 11:
                            pc_index += 1
                            cd_data.append({'C': pc_Data[pc_index-1]})
                    print("Moving data to Run 2")
                    
                    cd_data.move_range("C37:C48", rows= -31, cols=0)
                    cd_data.move_range("C29:C36", rows= -19, cols=0)
                    #cd.save("combined_Data.xlsx")

                # Column for Run 3 writing
                if check_data_c[6].value != None:
                    print("Populating Run 3")
                    for pc_r3 in pc_Data:
                        if pc_index <= 11:
                            pc_index += 1
                            cd_data.append({'D': pc_Data[pc_index-1]})
                    print("Moving data to Run 3")

                    cd_data.move_range("D25:D36", rows=-19, cols=0)
                    cd_data.move_range("D37:D40", rows= -31, cols=0)
                    #cd.save("combined_Data.xlsx")
        cd.save("combined_Data.xlsx")

    else:
        print("Benchmark data does not exist.")
    


def main():

    file_exists = exists("combined_Data.xlsx")

    if file_exists == True:
        cd = load_workbook(filename= "combined_Data.xlsx")
        cd_data = cd.active

    if file_exists == False:
        cd = Workbook()
        cd_data = cd.active
        # Setup the headers in a dict
        headers_a = (
            ('[Insert DUT]', None), ('Power Slider Mode', None, '[Insert AC/DC]'), ('EPP (for reference)', None), 
            ('[Insert DUT Info]', None), ('Run', 'R1', 'R2', 'R3'), ('PCMark10 Score', None), ('Essentials', None), 
            ('Productivity', None), ('Dig Content Creation', None), ('App Startup', None), ('Video Confrencing', None),
            ('Web Browsing', None), ('Spreadsheet', None), ('Writing', None), ('Photo Editing', None), 
            ('Render and Visual', None), ('Video Editing', None), ('DAQ MCP Power', None), ('Perf/Watt', None), 
            ('Crossmark Score', None), ('Productivity', None), ('Creativity', None), ('Responsiveness', None), 
            ('DAQ MCP Power', None))

        # Go through each dict entry and add it to the worksheet
        for i in headers_a:
            cd_data.append(i)

        # Fix up alignment
        cd_data['C4'].alignment = Alignment(horizontal='center')
        cd_data.column_dimensions['A'].width = 17

        int = 0

        # Set the colours and make it magical
        for cell in cd_data['A']:
            if cell.value != None:
                int = int+1

                if (6<=int<=9):
                    cd_data['A' + str(int)].fill = PatternFill('solid', start_color="00FF9900")

                if (20<=int<=20):
                    cd_data['A' + str(int)].fill = PatternFill('solid', start_color="00FF9900")

                if (21<=int<=23):
                    cd_data['A' + str(int)].fill = PatternFill('solid', start_color="00FFCC99")

                if (10<=int<=17):
                    cd_data['A' + str(int)].fill = PatternFill('solid', start_color="00FFCC99")
                
                if (18<=int<=19):
                    cd_data['A' + str(int)].fill = PatternFill('solid', start_color="00C0C0C0")

    # Make template
    template(cd_data, cd)

    # Save the workbook
    cd.save("combined_Data.xlsx")




if __name__ == "__main__":
    main()